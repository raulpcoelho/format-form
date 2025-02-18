from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import numbers


def format_sheet(file_path: str) -> None:
    workbook = load_workbook(file_path)
    sheet = workbook.active

    fraction_format = "# ?/?"
    integer_format = "0"
    center_alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = integer_format
            if isinstance(cell.value, (float)):
                cell.number_format = fraction_format
            cell.alignment = center_alignment

    workbook.save(file_path)
    print(f"File saved as {file_path}")
